import timeit as ti

def rust_test():
    import task

    f = open("input.xlsx", "rb")

    s = f.read()

    task.parse(s)

def oldpy_test():
    import json
    from xlrd import open_workbook

    wb = open_workbook('input.xlsx')

    sh = wb.sheet_by_index(0)

    values = []

    for r in sh.get_rows():
        item = {"sku": r[0].value, "designation": r[1].value, "price": r[2].value}
        item["description"] = "%s (%s), $%s" % (item["designation"], item["sku"], item["price"])
        values.append(item)

    json.dumps(values, ensure_ascii=False)


def py_test():
    import json
    from openpyxl import load_workbook
    wb = load_workbook(filename='input.xlsx')["Sheet1"]

    values = []

    for r in wb.iter_rows():
        item = {"sku": r[0].value, "designation": r[1].value, "price": r[2].value}
        item["description"] = "%s (%s), $%s" % (item["designation"], item["sku"], item["price"])
        values.append(item)

    json.dumps(values, ensure_ascii=False)


if __name__ == "__main__":
    setup = "from __main__ import rust_test, py_test, oldpy_test"
    
    print("Rust implementation: " + str(ti.timeit("rust_test()", setup=setup, number=10000)))
    print("openpyxl implementation: " + str(ti.timeit("py_test()", setup=setup, number=10000)))
    print("Deceprecated xlrd implementation: " + str(ti.timeit("oldpy_test()", setup=setup, number=10000)))
